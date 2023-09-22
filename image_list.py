import os
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, simpledialog
import time
import tkinter.messagebox
from PIL import Image, ImageTk,  ImageOps, UnidentifiedImageError
import piexif
from pathlib2 import Path
import shutil
from tkinterdnd2 import DND_FILES, TkinterDnD
import tempfile
from geopandas import GeoDataFrame
from shapely.geometry import Point

image_label = None

global listbox

def show_error_in_listbox(error_message, listbox):
    listbox.insert(tk.END, f"ERROR: {error_message}")



def create_widgets(root):
    try:
        global image_label

        # Create a frame for the treeview
        tree_frame = ttk.Frame(root, padding="10")
        tree_frame.grid(row=0, column=0, sticky='nsew')

        # Create the treeview inside the frame
        tree = ttk.Treeview(tree_frame,padding="10")
        tree.grid(row=0, column=0, sticky='nsew')

        # Create a frame for the image label
        image_frame = ttk.Frame(root, padding="10")
        image_frame.grid(row=1, column=1, sticky='nsew')

        # Create the image label inside the frame
        image_label = tk.Label(image_frame)
        image_label.pack()
    except Exception as e:
        show_error_in_listbox(f"Error in create_widgets: {e}",listbox)

def show_image_preview(tree, listbox):
    try:
        global image_label
        selected_items = tree.selection()  # Get the ID of the selected item
        if selected_items:  # Check if there is a selected item
            selected_item_id = selected_items[0]
            file_path = tree.set(selected_item_id, "fullpath")  # Get the full path of the selected item

            _, ext = os.path.splitext(file_path)
            if ext.lower() in ['.jpeg', '.jpg']:
                # Open and resize the image
                image = Image.open(file_path)
                image = image.resize((200, 200), Image.LANCZOS)  # Resize the image to 100x100 pixels

                # Create a PhotoImage object and set it as the image option of the image label
                photo_image = ImageTk.PhotoImage(image)
                image_label.config(image=photo_image)

                # Store the PhotoImage object as an attribute of the label to prevent it from being garbage collected
                image_label.photo_image = photo_image
    except UnidentifiedImageError:
        show_error_in_listbox("Invalid image format.",listbox)
    except Exception as e:
        show_error_in_listbox(f"Error in show_image_preview: {e}",listbox)

def populate_tree(tree, node):
    try:
        path = Path(tree.set(node, "fullpath"))
        if path.is_dir():
            for p in path.iterdir():
                is_dir = p.is_dir()
                oid = tree.insert(node, "end", text=p.name, values=[p, 'directory' if is_dir else 'file'])
                if is_dir:
                    tree.after(10, populate_tree, tree, oid)
    except Exception as e:
        show_error_in_listbox(f"Error in populate_tree: {e}",listbox)

def get_exif_data(image_path):
    try:
        exif_dict = piexif.load(image_path)
        return exif_dict
    except Exception as e:
        show_error_in_listbox(f"Failed to get EXIF data: {e}",listbox)

def get_coordinates(gps_info):
    try:
        lat_data = gps_info[piexif.GPSIFD.GPSLatitude]
        lon_data = gps_info[piexif.GPSIFD.GPSLongitude]

        # Convert the GPS coordinates stored in the EXIF to dd format
        lat_degree = lat_data[0][0] / lat_data[0][1]
        lat_minute = lat_data[1][0] / lat_data[1][1]
        lat_second = lat_data[2][0] / lat_data[2][1]
        lon_degree = lon_data[0][0] / lon_data[0][1]
        lon_minute = lon_data[1][0] / lon_data[1][1]
        lon_second = lon_data[2][0] / lon_data[2][1]

        latitude = lat_degree + (lat_minute / 60) + (lat_second / 3600)
        longitude = lon_degree + (lon_minute / 60) + (lon_second / 3600)

        # Adjust the sign of the DD values based on the hemisphere
        if gps_info[piexif.GPSIFD.GPSLatitudeRef] == b'S':
            latitude = -latitude
        if gps_info[piexif.GPSIFD.GPSLongitudeRef] == b'W':
            longitude = -longitude

        return latitude, longitude
    except Exception as e:
        show_error_in_listbox(f"Failed to get coordinates data: {e}", listbox)


def degrees_to_direction(degrees):
    try:
        directions = ['N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW', 'N']
        index = round(degrees / 45)
        return directions[index]
    except Exception as e:
        show_error_in_listbox(f"Failed to get degree direction data: {e}", listbox)
output_geojson_file = None
output_shapefile_dir = None

def process_images(directory, excel_path, progress_var, total_files, root, time_label, file_count_label, listbox, progress_label):
    if excel_path is None:
        excel_path = tempfile.mktemp(suffix=".xlsx")
    image_data_list = []
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Filename', 'Latitude', 'Longitude', 'DateTime', 'Orientation', 'Folder'])
        row = 2
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = 20
    print(f'Total files: {total_files}')
    start_time = time.time()
    processed_files = 0  # Initialize a counter for the processed files
    for dirpath, dirs, files in os.walk(directory):
        for filename in files:
            if filename.endswith((".JPG", ".jpg", ".jpeg")):
                try:
                    full_path = os.path.join(dirpath, filename)
                    folder_name = os.path.basename(os.path.dirname(full_path))

                    exif_data = get_exif_data(full_path)
                    # Print the entire EXIF dictionary in a sorted order
                    for ifd in ("0th", "Exif", "GPS", "1st"):
                        for tag in sorted(exif_data[ifd]):
                            tag_name = piexif.TAGS[ifd][tag]["name"]
                            value = exif_data[ifd][tag]
                            print(f"{tag_name}: {value}")

                    # Get the image direction in degrees
                    if piexif.GPSIFD.GPSImgDirection in exif_data['GPS']:
                        img_direction = exif_data['GPS'][piexif.GPSIFD.GPSImgDirection]
                        img_direction = img_direction[0] / img_direction[1]  # Get the actual value
                        # Convert the image direction to a compass direction
                        compass_direction = degrees_to_direction(img_direction)
                    else:
                        compass_direction = 'N/A'

                    # Get the GPS info
                    try:
                        gps_info = exif_data['GPS']
                        if gps_info:
                            latitude, longitude = get_coordinates(gps_info)
                        else:
                            latitude = 'N/A'
                            longitude = 'N/A'
                    except Exception as e:
                        show_error_in_listbox(f'Error getting GPS info: {e}', listbox)
                        latitude = 'N/A'
                        longitude = 'N/A'
                    datetime = exif_data.get('DateTime', '')
                    image = Image.open(full_path)
                    thumbnail = ImageOps.exif_transpose(image)
                    thumbnail.thumbnail((100, 100))
                    thumbnail_dir = os.path.join(directory, 'thumbnail')
                    os.makedirs(thumbnail_dir, exist_ok=True)

                    thumbnail_path = os.path.join(thumbnail_dir, f"{filename}_thumbnail.jpg")
                    thumbnail.save(thumbnail_path)

                    image_data_list.append({
                        'Filename': filename,
                        'Latitude': latitude,
                        'Longitude': longitude,
                        'DateTime': datetime,
                        'Orientation': compass_direction,
                        'Folder': folder_name
                    })

                    ws.append([filename, latitude, longitude, datetime, compass_direction, folder_name])

                    img = XLImage(thumbnail_path)
                    img.width = img.width * 1.5
                    img.height = img.height * 1.5
                    ws.add_image(img, f"G{row}")
                    ws.column_dimensions['G'].width = img.width // 7
                    ws.row_dimensions[row].height = img.height

                    row += 1
                    processed_files += 1  # Increment the counter for the processed files
                    elapsed_time = time.time() - start_time
                    remaining_files = total_files - processed_files
                    remaining_time = elapsed_time / processed_files * remaining_files
                    time_label['text'] = f"Estimated time remaining: {remaining_time:.2f} seconds"
                    file_count_label['text'] = f"Files remaining: {remaining_files}"
                    listbox.insert(tk.END, f"Processed file {filename} - {latitude}, {longitude} - {compass_direction}")
                    root.update()

                    progress_var.set(processed_files)  # Update the progress variable after the image has been processed
                    progress_percent = (processed_files / total_files) * 100  # Calculate the progress percentage
                    progress_label['text'] = f"{progress_percent:.2f}%"  # Update the progress label

                except Exception as e:
                    show_error_in_listbox(f"Error processing file {filename}: {e}", listbox)


    # Filtra le immagini che hanno le coordinate GPS valide
    geo_data = [d for d in image_data_list if d['Latitude'] != 'N/A' and d['Longitude'] != 'N/A']

    # Crea un GeoDataFrame
    geometry = [Point(xy) for xy in zip([float(d['Longitude']) for d in geo_data], [float(d['Latitude']) for d in geo_data])]
    geo_df = GeoDataFrame(geo_data, geometry=geometry)

    # Salva in GeoJSON
    if output_geojson_file:  # Se output_geojson_file è stato impostato
        geo_df.to_file(output_geojson_file, driver='GeoJSON')

    # Salva in shapefile
    if output_shapefile_dir:  # Se output_shapefile_dir è stato impostato
        geo_df.to_file(output_shapefile_dir, driver='ESRI Shapefile')

    wb.save(excel_path)
    tkinter.messagebox.showinfo("Information", "Finished processing images.")

input_dir = None
output_file = None
def drop(event, tree, root):  # Add root as an argument
    print("Drop event triggered.")  # Print a message when the drop event is triggered
    global input_dir
    files = root.tk.splitlist(event.data)  # Use splitlist() to get the list of dropped files
    for file in files:
        _, ext = os.path.splitext(file)
        if ext.lower() in ['.jpeg', '.jpg']:
            print(f"Source file: {file}")

            # Add the file to the tree
            tree.insert('', 'end', text=file, values=[file, "file"])

            # Copy the file to the selected directory in the tree
            selected_items = tree.selection()  # Get the ID of the selected item
            if selected_items:  # Check if there is a selected item
                selected_item_id = selected_items[0]
                input_dir = tree.set(selected_item_id, "fullpath")  # Get the full path of the selected item

                if os.path.isdir(input_dir):  # Make sure the destination is a directory
                    print(f"Destination directory: {input_dir}")
                    shutil.copy(file, input_dir)
                    root.update_idletasks()  # Update the UI
                    # Update the tree
                    tree.delete(
                        *tree.get_children(selected_item_id))  # Delete the current children of the selected item
                    populate_tree(tree, selected_item_id)  # Repopulate the tree with the new directory structure

                else:
                    show_error_in_listbox(f"{input_dir} is not a directory.",   listbox)
            else:
                show_error_in_listbox("No directory selected in the tree.", listbox)
def create_directories(tree):
    global input_dir  # Dichiarazione di input_dir come variabile globale

    base_dir = filedialog.askdirectory()  # Ask the user to choose a base directory
    prefix = simpledialog.askstring("Input", "Enter the prefix for directories:")  # Ask the user to input a prefix

    input_dir = base_dir  # Imposta input_dir sul base_dir

    range_str = simpledialog.askstring("Input", "Enter the range of directories to create (e.g., '1-10'):")
    ranges = range_str.split(';')
    for range_str in ranges:
        start, end = map(int, range_str.split('-'))

        for i in range(start, end + 1):
            dir_name = "{}{:05d}".format(prefix, i)  # This will create names like N00001, N00002, etc.
            dir_path = os.path.join(base_dir, dir_name)
            os.mkdir(dir_path)

    # Clear the tree before populating it
    for i in tree.get_children():
        tree.delete(i)

    # Populate the tree with the new directory structure
    root_node = tree.insert('', 'end', text=base_dir, values=[base_dir, "directory"])
    populate_tree(tree, root_node)
    if input_dir and output_file:  # Se sia input_dir che output_file sono impostati
        start_button['state'] = 'normal'  # Abilita il pulsante di avvio
def import_images(progress_var, time_label, file_count_label, listbox, root, progress_bar, tree):
    try:
        global input_dir  # Dichiarazione di input_dir come variabile globale
        input_dir = filedialog.askdirectory(title='Seleziona la cartella delle immagini')
        total_files = sum(len(files) for dirpath, dirs, files in os.walk(input_dir)
                          if any(file.endswith((".JPG", ".jpg", ".jpeg")) for file in files))  # Ottieni il numero totale di file immagine
        progress_var.set(0)  # Inizializza la variabile di progresso
        progress_bar["maximum"] = total_files  # Imposta il valore massimo della barra di avanzamento
        if input_dir and output_file:  # Se sia input_dir che output_file sono impostati
            start_button['state'] = 'normal'  # Abilita il pulsante di avvio

        # Pulisci l'albero prima di popolarlo
        for i in tree.get_children():
            tree.delete(i)

        # Popola l'albero con la struttura della directory
        root_node = tree.insert('', 'end', text=input_dir, values=[input_dir, "directory"])
        populate_tree(tree, root_node)
    except Exception as e:
        show_error_in_listbox(f"Failed to get import image data: {e}",  listbox)

def save_excel():
    try:
        global output_file # Declare output_file as a global variable
        output_file = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')], title='Save Excel File As')
        if input_dir and output_file:  # If both input_dir and output_file are set
            start_button['state'] = 'normal'  # Enable the start button
    except Exception as e:
        show_error_in_listbox(f"Failed to get excel data: {e}", listbox)
def save_geojson():
    try:
        global output_geojson_file  # Declare output_geojson_file as a global variable
        output_geojson_file = filedialog.asksaveasfilename(defaultextension='.geojson', filetypes=[('GeoJSON Files', '*.geojson')], title='Save GeoJSON File As')
        if input_dir and output_geojson_file:  # If both input_dir and output_geojson_file are set
            start_button['state'] = 'normal'  # Enable the start button
    except Exception as e:
        show_error_in_listbox(f"Failed to get GeoJSON data: {e}", listbox)

def save_shapefile():
    try:
        global output_shapefile_dir  # Declare output_shapefile_dir as a global variable
        output_shapefile_dir = filedialog.askdirectory(title='Save Shapefile Folder As')
        if input_dir and output_shapefile_dir:  # If both input_dir and output_shapefile_dir are set
            start_button['state'] = 'normal'  # Enable the start button
    except Exception as e:
        show_error_in_listbox(f"Failed to get shapefile data: {e}", listbox)
def start_processing(progress_var, time_label, file_count_label, listbox, root, progress_label):
    total_files = sum(len(files) for _, _, files in os.walk(input_dir))  # Get the total number of files
    progress_var.set(0)  # Initialize the progress variable
    process_images(input_dir, output_file, progress_var, total_files, root, time_label, file_count_label, listbox, progress_label)
def autoscroll(sbar, first, last):
    """Hide and show scrollbar as needed."""
    first, last = float(first), float(last)
    if first <= 0 and last >= 1:
        sbar.grid_remove()
    else:
        sbar.grid()
    sbar.set(first, last)

selected_item = None  # Variabile globale per tenere traccia dell'elemento selezionato

def on_item_select(event):
    global selected_item
    tree = event.widget
    selected_items = tree.selection()
    if not selected_items:  # Se la selezione è vuota
        return  # Esci dalla funzione
    selected_item = selected_items[0]
    print(f"Item selected: {selected_item}")

def on_item_drop(event, widget=None):
    global selected_item
    src_path = ''
    dest_dir = ''
    tree = event.widget
    target_item = tree.identify('item', event.x, event.y)
    if target_item and selected_item:
        if selected_item and tree.exists(selected_item):
            src_path = tree.item(selected_item)['values'][0]
            dest_dir = tree.item(target_item)['values'][0]

        # Verifica se src_path è un'immagine
    #if src_path and src_path.lower().endswith(('.jpg', '.jpeg')):
        if src_path.lower().endswith(('.jpg', '.jpeg')):
            if os.path.isdir(dest_dir):  # Assicurati che la destinazione sia una directory
                try:
                    # Esegui lo spostamento o la copia qui
                    shutil.move(src_path, os.path.join(dest_dir, os.path.basename(src_path)))
                    print(f"Spostato {src_path} in {dest_dir}")

                    # Trova il nodo padre per aggiornare solo quella parte dell'albero
                    parent_item = tree.parent(target_item)
                    update_tree(tree, parent_item)

                except Exception as e:
                    print(f"Errore durante lo spostamento: {e}")
            else:
                print(f"{dest_dir} non è una directory.")
        else:
            print("Il file trascinato non è un'immagine supportata.")




def update_tree(tree, parent_item):
    # Eliminare tutti i nodi figli
    tree.delete(*tree.get_children(parent_item))

    # Reinserire i nodi
    populate_tree(tree, parent_item)

def main():
    global input_dir, output_file, start_button  # Declare input_dir, output_file, and start_button as global variables
    #root = tk.Tk()
    root = TkinterDnD.Tk()  # Define root2 here
    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', lambda event: drop(event, tree,root))  # Remove root2 from the lambda function

    vsb = ttk.Scrollbar(orient="vertical")
    hsb = ttk.Scrollbar(orient="horizontal")

    tree = ttk.Treeview(columns=("fullpath", "type"), displaycolumns="")
                        #yscrollcommand=lambda f, l: autoscroll(vsb, f, l),
                        #xscrollcommand=lambda f, l: autoscroll(hsb, f, l))

    create_widgets(root)
    # Bind the <<TreeviewSelect>> event to the show_image_preview function
    tree.bind('<<TreeviewSelect>>', lambda event: show_image_preview(tree, listbox))


    tree.bind('<ButtonRelease-1>', lambda event, widget=None: on_item_drop(event))

    tree.bind('<Button-1>', on_item_select)  # Aggiungi questo bind per gestire la pressione del pulsante del mouse

    #vsb['command'] = tree.yview
    hsb['command'] = tree.xview

    tree.heading("#0", text="Directory Structure", anchor='w')
    # root_directory = Path('C:/')  # change this to your directory
    # root_node = tree.insert('', 'end', text=root_directory, values=[root_directory, "directory"])
    # populate_tree(tree, root_node)

    tree.grid(column=0, row=1, sticky='nsew')
    #vsb.grid(column=1, row=0, sticky='ns')
    hsb.grid(column=0, row=2, sticky='ew')

    root.grid_columnconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)  # Change this to 1
    root.title("Image Processor")

    menubar = tk.Menu(root)
    filemenu = tk.Menu(menubar, tearoff=0)
    filemenu.add_command(label="Import photos folder", command=lambda: import_images(progress_var, time_label, file_count_label, listbox, root, progress_bar, tree))

    # Aggiungi un sottomenu per l'esportazione vettoriale
    export_vector_menu = tk.Menu(filemenu, tearoff=0)
    export_vector_menu.add_command(label="Save GeoJSON", command=save_geojson)
    export_vector_menu.add_command(label="Save Shapefile", command=save_shapefile)

    filemenu.add_cascade(label="Vector Export", menu=export_vector_menu)  # Aggiunge il sottomenu al menu File
    filemenu.add_command(label="Save Excel", command=save_excel)
    filemenu.add_command(label="Create Directories", command=lambda : create_directories(tree))

    menubar.add_cascade(label="File", menu=filemenu)
    root.config(menu=menubar)



    # Create a frame to hold the listbox and scrollbar
    frame = tk.Frame(root)
    frame.grid(row=0, column=0, sticky='nsew')

    listbox = tk.Listbox(frame)  # Create a listbox to show the process in real time
    listbox.grid(row=0, column=0, sticky='nsew')

    scrollbar = tk.Scrollbar(frame, orient="vertical", command=listbox.yview)
    scrollbar.grid(row=0, column=1, sticky='ns')

    listbox.configure(yscrollcommand=scrollbar.set)
    progress_var = tk.DoubleVar()

    progress_frame = tk.Frame(root)  # Create a frame to hold the progress bar and label
    progress_frame.grid(row=3, column=0, sticky='ew')

    progress_bar = ttk.Progressbar(progress_frame, length=600, variable=progress_var)
    progress_bar.pack(fill='x')

    progress_label = tk.Label(progress_frame, text="0.00%")  # Create a label to show the progress percentage
    progress_label.place(relx=0.5, rely=0.5, anchor='center')  # Place the label at the center of the progress bar


    #progress_label.grid(row=2, column=1, sticky='w')
    start_button = tk.Button(root, text="Start", state='disabled',
                             command=lambda: start_processing(progress_var, time_label, file_count_label, listbox, root, progress_label))  # Create a label to show the progress percentage

    start_button.grid(row=2, column=0, sticky='ew')
    time_label = tk.Label(root)  # Create a label to show the estimated time remaining
    time_label.grid(row=3, column=0, sticky='w')
    file_count_label = tk.Label(root)  # Create a label to show the file count
    file_count_label.grid(row=5, column=0, sticky='w')

    # Configure the grid to expand properly when the window is resized
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)
    frame.grid_rowconfigure(1, weight=1)
    frame.grid_columnconfigure(0, weight=1)
    root.mainloop()
if __name__ == '__main__':
    main()


